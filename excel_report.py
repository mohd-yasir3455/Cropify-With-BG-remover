"""Excel report writer for ``processing_report.xlsx``.

Maintains one row per product folder with columns: ``Folder Name``,
``No. of Pics`` and ``Any Comment``. On a rerun the existing workbook is loaded
and rows are appended or updated in place -- the sheet is never rebuilt from
scratch, so prior results and manual notes survive. A folder that appears again
updates its own row instead of creating a duplicate.

Saving is atomic (temp file + replace) to avoid corrupting the report if the
process is interrupted mid-write.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
from logger import get_logger

logger = get_logger()

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(name=config.EXCEL_FONT, bold=True, color="FFFFFF")
_BODY_FONT = Font(name=config.EXCEL_FONT)
_COLUMN_WIDTHS = (30, 12, 62)


class ExcelReport:
    """Append/update-only report backed by an ``openpyxl`` workbook."""

    def __init__(self, path: Path) -> None:
        """Open the report at ``path`` if it exists, else create a new one."""
        self.path = path
        self._row_for: dict[str, int] = {}

        if path.exists():
            try:
                self.wb = load_workbook(path)
                self.ws = self.wb.active
                self._index_existing()
            except Exception:  # corrupt/locked file -> start a fresh workbook
                logger.exception("Could not open existing report; creating a new one.")
                self._new_workbook()
        else:
            self._new_workbook()

        self._next_row = max(self.ws.max_row + 1, 2)

    # -- construction helpers --------------------------------------------- #
    def _new_workbook(self) -> None:
        """Initialise a fresh workbook with a styled header row."""
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Report"
        self._write_header()

    def _write_header(self) -> None:
        """Write and style the three-column header, freeze it, size columns."""
        for col, title in enumerate(config.EXCEL_HEADERS, start=1):
            cell = self.ws.cell(row=1, column=col, value=title)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        self.ws.freeze_panes = "A2"
        for col, width in enumerate(_COLUMN_WIDTHS, start=1):
            self.ws.column_dimensions[get_column_letter(col)].width = width

    def _index_existing(self) -> None:
        """Map existing folder names (column A) to their row numbers."""
        if self.ws.cell(row=1, column=1).value != config.EXCEL_HEADERS[0]:
            # Older/foreign file without our header: add one at the top.
            self.ws.insert_rows(1)
            self._write_header()
        for row in range(2, self.ws.max_row + 1):
            name = self.ws.cell(row=row, column=1).value
            if name is not None:
                self._row_for[str(name)] = row

    # -- public API -------------------------------------------------------- #
    def upsert(self, folder: str, count: int, comment: str) -> None:
        """Insert or update the row for ``folder``.

        Args:
            folder: Product folder name (unique key, column A).
            count: Number of valid images in the folder.
            comment: Text for the "Any Comment" column (may be empty).
        """
        row = self._row_for.get(folder)
        if row is None:
            row = self._next_row
            self._next_row += 1
            self._row_for[folder] = row

        name_cell = self.ws.cell(row=row, column=1, value=folder)
        name_cell.font = _BODY_FONT
        count_cell = self.ws.cell(row=row, column=2, value=count)
        count_cell.font = _BODY_FONT
        count_cell.alignment = Alignment(horizontal="center")
        comment_cell = self.ws.cell(row=row, column=3, value=comment)
        comment_cell.font = _BODY_FONT
        comment_cell.alignment = Alignment(wrap_text=True, vertical="top")

    def save(self) -> None:
        """Atomically persist the workbook to disk."""
        self.path.parent.mkdir(parents=True, exist_ok=True)
        tmp = self.path.with_suffix(".xlsx.tmp")
        self.wb.save(tmp)
        tmp.replace(self.path)
